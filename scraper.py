"""
Dynamic Business Scraper 💎
A Beta lead generation tool that scrapes Google Maps, BBB, and Yellow Pages.
"""

import time
import re
import sys
import requests
from typing import List, Optional, Dict, Any
from urllib.parse import urljoin, quote

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class DynamicProScraper:
    """
    A scraper class to find business leads and their owners.
    """

    def __init__(self, niche: str):
        """
        Initialize the scraper with a target niche.
        
        Args:
            niche: The business category to search for (e.g., 'Solar').
        """
        self.niche = niche
        clean_name = re.sub(r'[^\w\s-]', '', niche).strip().replace(' ', '_')
        self.filename = f"{clean_name}.xlsx"
        self.data: List[Dict[str, Any]] = []
        self.driver: Optional[webdriver.Chrome] = None
        self.owner_patterns = [
            r'(?:Owner|President|CEO|Founder|Principal)[:\s.,]*([A-Z][a-z]+\s[A-Z][a-z]+(?:\s[A-Z][a-z]+)?)',
            r'([A-Z][a-z]+\s[A-Z][a-z]+(?:\s[A-Z][a-z]+)?),\s*(?:Owner|President|CEO|Founder|Principal)',
            r'(?:Founded?\s+by|Established\s+by)[:\s.,]*([A-Z][a-z]+\s[A-Z][a-z]+)',
            r'([A-Z][a-z]{2,}\s[A-Z][a-z]{2,})\s+(?:Owner|President|CEO)',
            r'Contact\s+([A-Z][a-z]+\s[A-Z][a-z]+)'
        ]
        self.setup_driver()

    def setup_driver(self) -> None:
        """Configures and initializes the Selenium WebDriver."""
        print("🔧 Initializing browser session...")
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument("--lang=en-US")
        chrome_options.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        prefs = {
            "profile.default_content_setting_values.notifications": 2, 
            "intl.accept_languages": "en-US,en"
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.driver.set_page_load_timeout(30)
        self.driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )

    def check_browser_open(self) -> bool:
        """
        Verify if the browser is still open.
        
        Returns:
            True if browser is accessible, False otherwise.
        """
        try:
            if self.driver:
                self.driver.current_window_handle
                return True
        except Exception:
            pass
        return False

    def search(self) -> None:
        """Executes search queries on Google Maps and processes results."""
        search_queries = [
            f"{self.niche} near USA", 
            f"{self.niche} repair USA", 
            f"best {self.niche} USA"
        ]
        
        for query in search_queries:
            if not self.check_browser_open(): 
                self.safe_exit("Browser was closed by user")
            
            print(f"\n📡 SEARCHING: {query}")
            try:
                self.driver.get(f"https://www.google.com/maps/search/{quote(query)}?hl=en")
                container = WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((
                        By.XPATH, 
                        "//div[contains(@role, 'feed')] | //div[contains(@aria-label, 'Results')]"
                    ))
                )
                
                # Scroll to load more results
                for _ in range(3):
                    if not self.check_browser_open(): 
                        self.safe_exit("Browser was closed by user")
                    self.driver.execute_script(
                        "arguments[0].scrollTop = arguments[0].scrollHeight", 
                        container
                    )
                    time.sleep(2)

                listings = self.driver.find_elements(By.XPATH, "//a[contains(@href, '/maps/place/')]")
                urls = list(dict.fromkeys([l.get_attribute('href') for l in listings]))[:12]
                
                for url in urls:
                    if not self.check_browser_open(): 
                        self.safe_exit("Browser was closed by user")
                    self.process_listing(url)
                    
            except WebDriverException as e:
                if any(err in str(e).lower() for err in ["disconnected", "target window already closed"]):
                    self.safe_exit("Browser connection lost")
                print(f"⚠️ Search error: {e}")

    def process_listing(self, url: str) -> None:
        """
        Extracts details from a single Google Maps listing.
        
        Args:
            url: The Google Maps URL of the business.
        """
        try:
            self.driver.get(url)
            time.sleep(3)
            name = self.driver.find_element(By.XPATH, "//h1").text.strip()
            
            # Base data extraction
            page_text = self.driver.find_element(By.TAG_NAME, 'body').text
            
            # Website extraction
            website = "N/A"
            try:
                website_elem = self.driver.find_element(
                    By.XPATH, "//a[contains(@data-item-id, 'authority')]"
                )
                website = website_elem.get_attribute('href')
            except Exception:
                web_match = re.search(
                    r'[a-zA-Z0-9.-]+\.(?:com|net|org|biz|us|co|io)', 
                    page_text.lower()
                )
                if web_match: 
                    website = "http://" + web_match.group(0)

            # Phone extraction
            phone = "N/A"
            phone_match = re.search(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', page_text)
            if phone_match: 
                phone = phone_match.group(0)

            # Address extraction
            city, state, address = "N/A", "N/A", "N/A"
            addr_elems = self.driver.find_elements(
                By.XPATH, "//button[contains(@aria-label, 'Address')]"
            )
            if addr_elems:
                address = addr_elems[0].text
                parts = address.split(',')
                if len(parts) >= 3:
                    city = parts[-3].strip()
                    st_parts = parts[-2].strip().split()
                    if st_parts: 
                        state = st_parts[0]

            # Skiptracing for owner
            owner = self.skiptrace_owner(name, website, city, state)
            
            self.data.append({
                'Company Name': name, 
                'Decision Maker': owner, 
                'Phone': phone, 
                'City': city, 
                'State': state, 
                'Website': website, 
                'Address': address
            })
            print(f"✓ {name} | Owner: {owner}")
            
        except Exception as e:
            if "target window already closed" in str(e).lower(): 
                self.safe_exit("Browser was closed")
            print(f"❌ Skipping listing: {e}")

    def skiptrace_owner(self, name: str, website: str, city: str, state: str) -> str:
        """
        Attempts to find the owner of a business using various sources.
        
        Returns:
            The name of the owner or 'N/A'.
        """
        if website and website != 'N/A':
            owner = self.get_owner_from_site_tab(website)
            if owner != 'N/A': 
                return owner
                
        if city and state:
            city_state = f"{city} {state}"
            owner = self.bbb_skiptrace(name, city_state)
            if owner: 
                return owner
            owner = self.yellowpages_skiptrace(name, city_state)
            if owner: 
                return owner
                
        return "N/A"

    def get_owner_from_site_tab(self, url: str) -> str:
        """Checks the company's own website for owner information."""
        main_tab = self.driver.current_window_handle
        owner = "N/A"
        try:
            self.driver.execute_script("window.open('');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            self.driver.get(url)
            time.sleep(3)
            
            # Look for common 'About' or 'Team' links
            links = self.driver.find_elements(By.TAG_NAME, "a")
            for l in links:
                try:
                    if any(k in l.text.lower() for k in ['about', 'team', 'meet', 'owner']):
                        self.driver.execute_script("arguments[0].click();", l)
                        time.sleep(2)
                        break
                except Exception: 
                    continue
                    
            text = self.driver.find_element(By.TAG_NAME, "body").text
            for p in self.owner_patterns:
                m = re.search(p, text)
                if m:
                    candidate = m.group(1).strip()
                    if len(candidate.split()) >= 2:
                        owner = candidate
                        break
        except Exception: 
            pass
        finally:
            try:
                if len(self.driver.window_handles) > 1: 
                    self.driver.close()
                self.driver.switch_to.window(main_tab)
            except Exception: 
                pass
        return owner

    def bbb_skiptrace(self, name: str, city_state: str) -> Optional[str]:
        """Queries the Better Business Bureau for business owner info."""
        try:
            url = f"https://www.bbb.org/search/?find_text={quote(name)}&find_loc={quote(city_state)}"
            resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=8)
            soup = BeautifulSoup(resp.text, 'html.parser')
            link = soup.select_one('a[href*="/profile/"] , a[href*="/usr/bbb/"]')
            if link:
                profile_url = urljoin("https://www.bbb.org", link['href'])
                profile_resp = requests.get(profile_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=8)
                for p in self.owner_patterns:
                    m = re.search(p, profile_resp.text)
                    if m: 
                        return m.group(1).strip()
        except Exception: 
            pass
        return None

    def yellowpages_skiptrace(self, name: str, city_state: str) -> Optional[str]:
        """Queries Yellow Pages for business owner info."""
        try:
            url = f"https://www.yellowpages.com/search?search_terms={quote(name)}&geo_location_terms={quote(city_state)}"
            resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=8)
            soup = BeautifulSoup(resp.text, 'html.parser')
            link = soup.select_one('a.business-name')
            if link:
                profile_url = urljoin("https://www.yellowpages.com", link['href'])
                profile_resp = requests.get(profile_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=8)
                for p in self.owner_patterns:
                    m = re.search(p, profile_resp.text)
                    if m: 
                        return m.group(1).strip()
        except Exception: 
            pass
        return None

    def save(self) -> None:
        """Saves collected data to a formatted Excel file."""
        if not self.data:
            print("No data to save.")
            return
            
        df = pd.DataFrame(self.data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        headers = list(df.columns)
        ws.append(headers)
        
        for r in self.data: 
            ws.append([r.get(h, "N/A") for h in headers])
            
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_len + 2, 50)
            
        wb.save(self.filename)
        print(f"\n📂 Leads saved to: {self.filename}")

    def safe_exit(self, reason: str = "User exit") -> None:
        """Saves progress and exits safely."""
        print(f"\n🛑 Aborting: {reason}. Saving results...")
        self.save()
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
        sys.exit(0)

def main():
    """Main entry point for the script."""
    print("💎 DYNAMIC BUSINESS SCRAPER")
    try:
        niche = input("🎯 Target Niche (e.g. Solar): ").strip() or "Leads"
        scraper = DynamicProScraper(niche)
        try:
            scraper.search()
        except (KeyboardInterrupt, WebDriverException):
            scraper.safe_exit("Manual stop or browser closed")
        finally:
            scraper.save()
            if scraper.driver:
                try:
                    scraper.driver.quit()
                except Exception:
                    pass
        print("🎬 Scrape complete.")
    except Exception as e:
        print(f"CRITICAL ERROR: {e}")

if __name__ == "__main__":
    main()
